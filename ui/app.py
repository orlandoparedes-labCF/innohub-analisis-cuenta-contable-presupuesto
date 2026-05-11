import streamlit as st
import pandas as pd
from groq import Groq
from dotenv import load_dotenv

load_dotenv()
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json
import re
import io
import os
from datetime import datetime

# ── Config ──────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Agente Contable | Comunidad Feliz",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

IVA = 1.19  # Chile IVA 19%

MONTHS_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo",  6: "Junio",   7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# ── Helpers ──────────────────────────────────────────────────────────────────
def fmt_clp(v):
    try:
        return f"${float(v):,.0f}"
    except Exception:
        return str(v)

def to_num(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)

def _border_thin(color="CCCCCC"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(color="AAAAAA"):
    return Border(bottom=Side(style="thin", color=color))


# ── 1. Parse libro mayor ─────────────────────────────────────────────────────
def parse_libro_mayor(file_bytes):
    df_raw = pd.read_excel(file_bytes, sheet_name=0, header=None, dtype=str)

    title   = str(df_raw.iloc[0, 0]) if pd.notna(df_raw.iloc[0, 0]) else ""
    company = str(df_raw.iloc[1, 0]) if pd.notna(df_raw.iloc[1, 0]) else ""

    saldo_anterior = 0
    for i in range(min(10, len(df_raw))):
        row_vals = [str(v) for v in df_raw.iloc[i] if pd.notna(v) and str(v) != "nan"]
        if any("saldo anterior" in v.lower() for v in row_vals):
            for v in row_vals:
                try:
                    saldo_anterior = float(v.replace(",", ".").replace(" ", ""))
                    break
                except ValueError:
                    pass

    # Find header row
    header_row_idx = 5
    for i in range(len(df_raw)):
        vals = " ".join(str(v) for v in df_raw.iloc[i] if pd.notna(v))
        if "Cuenta" in vals and "Fecha" in vals:
            header_row_idx = i
            break

    raw_headers = df_raw.iloc[header_row_idx].tolist()
    headers = [str(h) if pd.notna(h) else f"Col_{i}" for i, h in enumerate(raw_headers)]

    # Deduplicate column names
    seen = {}
    for i, h in enumerate(headers):
        if h in seen:
            seen[h] += 1
            headers[i] = f"{h} {seen[h]}"
        else:
            seen[h] = 1

    data_rows = []
    for i in range(header_row_idx + 1, len(df_raw)):
        row     = df_raw.iloc[i]
        row_str = " ".join(str(v) for v in row if pd.notna(v) and str(v) != "nan")
        if row_str.strip() == "":
            continue
        if "Subtotal" in row_str or row_str.strip().startswith("Total"):
            break
        non_null = [v for v in row if pd.notna(v) and str(v).strip() not in ("nan", "")]
        if len(non_null) >= 3:
            data_rows.append(row.tolist())

    if not data_rows:
        raise ValueError("No se encontraron filas de datos en el archivo.")

    df = pd.DataFrame(data_rows, columns=headers)

    # Extract account name from column 0
    account_name = ""
    if len(df) > 0:
        val = str(df.iloc[0, 0])
        if val not in ("nan", "", "None"):
            account_name = val

    # Detect periods from date column (col index 1)
    periods = []
    period_totals = {}
    date_col = df.columns[1] if len(df.columns) > 1 else None
    deb_col  = df.columns[3] if len(df.columns) > 3 else None
    cre_col  = df.columns[4] if len(df.columns) > 4 else None

    if date_col:
        dates = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce")
        valid = dates.dropna()
        if len(valid) > 0:
            df["__period"] = dates.dt.to_period("M")
            unique_periods = sorted(df["__period"].dropna().unique())
            periods = [str(p) for p in unique_periods]

            if deb_col and cre_col:
                df["__deb_num"] = to_num(df[deb_col])
                df["__cre_num"] = to_num(df[cre_col])
                for p in unique_periods:
                    mask = df["__period"] == p
                    period_totals[str(p)] = {
                        "deb": float(df.loc[mask, "__deb_num"].sum()),
                        "cre": float(df.loc[mask, "__cre_num"].sum()),
                    }

    return df, title, company, saldo_anterior, account_name, periods, period_totals


# ── 2. AI analysis ───────────────────────────────────────────────────────────
def _parse_json_safe(text: str) -> dict:
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        text = text.replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
        return json.loads(text)


def get_ai_analysis(providers: list, tx_summary: str, account_name: str,
                    api_key: str, periods: list) -> dict:
    client = Groq(api_key=api_key)

    multi_period = len(periods) > 1
    period_schema = ""
    if multi_period:
        period_schema = """,
  "analisis_por_periodo": [
    {
      "periodo": "YYYY-MM (ej: 2024-03)",
      "mes_nombre": "nombre del mes en español",
      "observacion": "En 2 oraciones: qué pasó en este mes específico, si fue un mes con gasto alto/bajo y por qué podría ser."
    }
  ]"""

    system = (
        "Eres un analista financiero experto en contabilidad chilena. "
        "Tu trabajo es analizar una cuenta contable específica del libro mayor de COMUNIDAD FELIZ SPA, "
        "empresa de administración de condominios en Chile. "
        "IMPORTANTE: El análisis es sobre una CUENTA CONTABLE ESPECÍFICA, no sobre la empresa completa. "
        "Enfócate solo en lo que ocurrió en esa cuenta durante el período analizado. "
        "Explica todo en lenguaje muy simple, como si se lo contaras a alguien de 15 años "
        "que nunca ha visto contabilidad. Usa lenguaje cotidiano, nada de tecnicismos. "
        "Los montos son en pesos chilenos (CLP)."
    )

    period_note = ""
    if multi_period:
        period_note = f"\nEsta cuenta tiene movimientos en {len(periods)} meses: {', '.join(periods)}."

    prompt = f"""Analiza la cuenta contable "{account_name}" del libro mayor de COMUNIDAD FELIZ SPA.{period_note}

PROVEEDORES QUE APARECEN EN ESTA CUENTA:
{json.dumps(providers, ensure_ascii=False, indent=2)}

RESUMEN DE MOVIMIENTOS DE ESTA CUENTA:
{tx_summary}

NOTA IMPORTANTE: Las filas marcadas como "Sin proveedor identificado" corresponden a transacciones
que fueron cargadas a esta cuenta pero que no tienen un proveedor específico registrado (por ejemplo,
distribuciones internas, reclasificaciones contables o asientos manuales). NO son gastos varios
generales de la empresa.

Responde ÚNICAMENTE con un JSON válido con esta estructura exacta:
{{
  "proveedores": [
    {{
      "rut": "RUT exacto tal como aparece en el archivo",
      "nombre": "nombre del proveedor",
      "descripcion_negocio": "En 2 oraciones simples: a qué se dedica esta empresa. Como si se lo explicaras a un adolescente que nunca ha oído de esta empresa.",
      "categoria": "Una etiqueta corta para clasificar este gasto (ej: Agencia de Marketing, Software, Publicidad Digital, Diseño Gráfico)",
      "relevancia_para_cf": "1 oración: para qué sirve este proveedor en el contexto específico de la cuenta '{account_name}'"
    }}
  ],
  "analisis_general": "Análisis de la cuenta '{account_name}' en 4 párrafos simples:\\nPárrafo 1: Explica en términos simples de qué trata esta cuenta contable y para qué sirve en una empresa de administración de condominios.\\nPárrafo 2: Cuáles fueron los proveedores más importantes por monto y qué representa cada uno.\\nPárrafo 3: Si hay notas de crédito (créditos/devoluciones), explica qué significan en términos simples.\\nPárrafo 4: Conclusión: si el nivel de gasto parece normal para este tipo de cuenta, si hay algo llamativo o inusual, y qué debería saber alguien que revisa este reporte por primera vez.",
  "alertas": [
    "Escribe entre 2 y 4 observaciones concretas sobre esta cuenta. Pueden ser positivas o de atención. Sé específico con montos y nombres."
  ],
  "resumen_ejecutivo": "En exactamente 3 oraciones muy simples: 1) Qué es esta cuenta y para qué se usa. 2) Cuánto se gastó en total y quién fue el proveedor principal. 3) Qué significa esto para Comunidad Feliz."{period_schema}
}}"""

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        messages=[
            {"role": "system", "content": system},
            {"role": "user",   "content": prompt},
        ],
        temperature=0.3,
        max_tokens=4096,
    )

    text = response.choices[0].message.content.strip()
    if "```json" in text:
        text = text.split("```json")[1].split("```")[0].strip()
    elif "```" in text:
        text = text.split("```")[1].split("```")[0].strip()

    return _parse_json_safe(text)


# ── 3. Generate Excel ────────────────────────────────────────────────────────
def generate_excel(df, analysis, title, company, saldo_ant,
                   account_name, periods, period_totals) -> bytes:
    wb  = openpyxl.Workbook()

    # ── PALETTE ──
    NAVY      = "1B3A6B"
    NAVY_MID  = "2C5282"
    WHITE     = "FFFFFF"
    LGRAY     = "F7F8FA"
    MGRAY     = "E2E6ED"
    DGRAY     = "404A5A"
    DARK      = "1A1A2E"
    MID       = "4A5568"
    RED       = "C0392B"
    GREEN_BG  = "EAF5EA"
    BLUE_BG   = "EBF4FF"
    AMBER_BG  = "FFF8E1"
    TEAL_BG   = "E0F4F4"
    GREEN_TX  = "1E7B34"
    BLUE_TX   = "1565C0"
    AMBER_TX  = "E65100"
    TEAL_TX   = "00695C"
    IVA_BG    = "FFF3E0"   # naranja suave para columna IVA
    IVA_TX    = "BF360C"   # naranja oscuro

    # ── COLORES ALTERNADOS PARA DISTRIBUCIÓN ──
    PERIOD_HDR = "37474F"   # cabecera de períodos (gris carbón)
    TOTAL_COL  = "263238"   # columna totales

    # ======== SHEET 1: MAYOR — texto plano ========
    ws1       = wb.active
    ws1.title = "Mayor"
    ncols     = len([c for c in df.columns if not c.startswith("__")])

    ws1.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws1["A1"].value     = f"{title}  |  {company}"
    ws1["A1"].font      = Font(name="Calibri", size=10, color="555555")
    ws1["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[1].height = 16

    ws1.merge_cells(f"A2:{get_column_letter(ncols)}2")
    ws1["A2"].value     = f"Cuenta: {account_name}  |  Saldo anterior: ${saldo_ant:,.0f}"
    ws1["A2"].font      = Font(name="Calibri", size=10, color="555555")
    ws1["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws1.row_dimensions[2].height = 16

    visible_cols = [c for c in df.columns if not c.startswith("__")]
    for ci, col in enumerate(visible_cols, 1):
        c           = ws1.cell(row=3, column=ci, value=str(col))
        c.font      = Font(bold=True, name="Calibri", size=10)
        c.fill      = PatternFill("solid", start_color="F2F2F2")
        c.border    = _border_thin()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[3].height = 28

    num_kw  = ["bit", "dito", "mulado"]
    date_kw = ["fecha"]

    for ri, (_, row) in enumerate(df.iterrows(), 4):
        for ci, col in enumerate(visible_cols, 1):
            val = row[col]
            c   = ws1.cell(row=ri, column=ci)
            sv  = str(val) if pd.notna(val) else ""
            if sv in ("nan", "None", "-", ""):
                c.value = ""
            else:
                try:
                    num     = float(sv.replace(",", ""))
                    c.value = int(num) if num == int(num) else num
                except (ValueError, OverflowError):
                    c.value = sv
            c.font      = Font(name="Calibri", size=10)
            c.border    = _border_thin()
            col_l = str(col).lower()
            if any(k in col_l for k in num_kw):
                c.number_format = "#,##0;(#,##0);\"-\""
                c.alignment     = Alignment(horizontal="right", vertical="center")
            elif any(k in col_l for k in date_kw):
                c.alignment     = Alignment(horizontal="center", vertical="center")
            else:
                c.alignment     = Alignment(horizontal="left", vertical="center")
        ws1.row_dimensions[ri].height = 15

    for i, w in enumerate([28,13,9,13,13,13,18,14,32,55,10,20,11,15,15][:ncols], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    ws1.freeze_panes = "A4"

    # ======== SHEET 2: ANÁLISIS — profesional ========
    ws2       = wb.create_sheet("Análisis")
    ws2.sheet_view.showGridLines = False

    # max cols usados en esta hoja (proveedor table = 9 cols A-I)
    MAX_COL = "I"

    def sec_title(row, text, height=22, bg=NAVY_MID, cols="A:I"):
        ws2.merge_cells(f"A{row}:{cols.split(':')[1]}{row}")
        c           = ws2[f"A{row}"]
        c.value     = text.upper()
        c.font      = Font(bold=True, name="Calibri", size=10, color=WHITE)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.row_dimensions[row].height = height

    def text_block(row, text, height=55, bg=WHITE, size=11, color=MID, cols="A:I"):
        ws2.merge_cells(f"A{row}:{cols.split(':')[1]}{row}")
        c           = ws2[f"A{row}"]
        c.value     = text
        c.font      = Font(name="Calibri", size=size, color=color)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(wrap_text=True, vertical="top", indent=1)
        ws2.row_dimensions[row].height = height

    def spacer(row, h=8):
        ws2.row_dimensions[row].height = h

    def kpi_card(row, col, label, value, bg, text_color):
        c = ws2.cell(row=row, column=col, value=label)
        c.font      = Font(name="Calibri", size=9, bold=True, color=text_color)
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="center", vertical="bottom")
        ws2.row_dimensions[row].height = 18

        c2 = ws2.cell(row=row+1, column=col, value=value)
        c2.font      = Font(name="Calibri", size=14, bold=True, color=text_color)
        c2.fill      = PatternFill("solid", start_color=bg)
        c2.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[row+1].height = 26

        for rr in [row, row+1]:
            ws2.cell(rr, col).border = Border(
                bottom=Side(style="medium", color=text_color) if rr == row+1 else Side(style=None),
                left=Side(style="thin", color="DDDDDD"),
                right=Side(style="thin", color="DDDDDD"),
                top=Side(style="thin", color="DDDDDD") if rr == row else Side(style=None),
            )

    r = 1

    # ── HEADER BLOCK (navy) ──
    ws2.merge_cells(f"A{r}:I{r}")
    c = ws2[f"A{r}"]
    c.value     = f"  {account_name}"
    c.font      = Font(bold=True, name="Calibri", size=16, color=WHITE)
    c.fill      = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[r].height = 36; r += 1

    ws2.merge_cells(f"A{r}:I{r}")
    c = ws2[f"A{r}"]
    c.value     = f"  {title}  ·  {company}  ·  Generado {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    c.font      = Font(name="Calibri", size=10, color="BED0E8")
    c.fill      = PatternFill("solid", start_color=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws2.row_dimensions[r].height = 18; r += 1

    spacer(r, 12); r += 1

    # ── KPI CARDS ──
    deb_col    = df.columns[3] if len(df.columns) > 3 else None
    cre_col    = df.columns[4] if len(df.columns) > 4 else None
    rut_col    = df.columns[7] if len(df.columns) > 7 else None
    nombre_col = df.columns[8] if len(df.columns) > 8 else None

    total_deb = to_num(df[deb_col]).sum() if deb_col else 0
    total_cre = to_num(df[cre_col]).sum() if cre_col else 0
    neto      = total_deb - total_cre
    n_tx      = len(df)
    n_prov    = df[rut_col].nunique() if rut_col else 0

    kpi_row = r
    kpi_card(kpi_row, 1, "💸 Total cargado",  fmt_clp(total_deb), BLUE_BG,  BLUE_TX)
    kpi_card(kpi_row, 3, "🔄 Total devuelto", fmt_clp(total_cre), GREEN_BG, GREEN_TX)
    kpi_card(kpi_row, 5, "📊 Saldo neto",     fmt_clp(neto),      AMBER_BG, AMBER_TX)
    kpi_card(kpi_row, 7, "📋 Movimientos",    str(n_tx),          TEAL_BG,  TEAL_TX)
    r += 3

    # ── NOTA IVA destacada ──
    ws2.merge_cells(f"A{r}:I{r}")
    c = ws2[f"A{r}"]
    c.value = (
        "  ⚠️  IMPORTANTE: Los montos de esta hoja se muestran SIN IVA (valores contables netos). "
        "La columna 'Neto + IVA' aplica el 19% SOLO a proveedores con RUT chileno (operan en Chile). "
        "Proveedores extranjeros / CLAY (sin RUT) quedan con su valor neto sin IVA."
    )
    c.font      = Font(name="Calibri", size=10, bold=True, color=IVA_TX)
    c.fill      = PatternFill("solid", start_color=IVA_BG)
    c.alignment = Alignment(wrap_text=True, vertical="center", indent=1)
    ws2.row_dimensions[r].height = 36; r += 1

    spacer(r, 10); r += 1

    # ── RESUMEN ──
    sec_title(r, "📌 Resumen ejecutivo"); r += 1
    text_block(r, analysis.get("resumen_ejecutivo", ""),
               height=60, bg=LGRAY, size=11, color=DARK); r += 1

    spacer(r); r += 1

    # ── OBSERVACIONES ──
    alertas = analysis.get("alertas", [])
    if alertas:
        sec_title(r, "⚡ Observaciones importantes"); r += 1
        for alerta in alertas:
            text_block(r, f"  ·  {alerta}", height=26, bg=WHITE, size=10, color=MID)
            r += 1

    spacer(r); r += 1

    # ── TABLA PROVEEDORES (9 columnas, col I = Neto + IVA) ──
    sec_title(r, "🏢 Detalle por proveedor  ·  valores contables sin IVA excepto última columna"); r += 1

    prov_cols = [
        "RUT",
        "Empresa / Proveedor",
        "¿A qué se dedica?",
        "Tipo de gasto",
        "Relación con esta cuenta",
        "Total cargado ($)",
        "Total abonado ($)",
        "Saldo neto ($)",
        "Neto + IVA 19% ($) ★",   # col I
    ]
    tbl_hdr_row = r
    for ci, h in enumerate(prov_cols, 1):
        c           = ws2.cell(row=tbl_hdr_row, column=ci, value=h)
        is_iva      = (ci == 9)
        c.font      = Font(bold=True, name="Calibri", size=10, color=WHITE)
        c.fill      = PatternFill("solid", start_color=IVA_TX if is_iva else DGRAY)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border_thin("AAAAAA")
    ws2.row_dimensions[tbl_hdr_row].height = 38; r += 1

    # Build provider groups
    ai_by_rut = {p["rut"]: p for p in analysis.get("proveedores", [])}

    df2             = df.copy()
    df2["__rut"]    = df2[rut_col].astype(str)    if rut_col    else ""
    df2["__nombre"] = df2[nombre_col].astype(str) if nombre_col else ""
    df2["__deb"]    = to_num(df2[deb_col])        if deb_col    else 0
    df2["__cre"]    = to_num(df2[cre_col])        if cre_col    else 0

    mask = ~df2["__rut"].isin(["", "nan", "-", "None"])
    grp  = df2[mask].groupby(["__rut", "__nombre"], as_index=False).agg(
        deb=("__deb", "sum"), cre=("__cre", "sum")
    )
    grp["neto"] = grp["deb"] - grp["cre"]
    grp = grp.sort_values("deb", ascending=False).reset_index(drop=True)

    df_no_rut = df2[~mask]
    if len(df_no_rut) > 0 and df_no_rut["__deb"].sum() > 0:
        gv = pd.DataFrame([{
            "__rut":    "-",
            "__nombre": "Transacciones sin proveedor identificado",
            "deb":  df_no_rut["__deb"].sum(),
            "cre":  df_no_rut["__cre"].sum(),
            "neto": df_no_rut["__deb"].sum() - df_no_rut["__cre"].sum(),
        }])
        grp = pd.concat([grp, gv], ignore_index=True)

    tbl_data_start = r
    for ri_p, row_p in grp.iterrows():
        row_num  = r + ri_p
        rut      = row_p["__rut"]
        nombre   = row_p["__nombre"]
        ai       = ai_by_rut.get(rut, {})
        bg       = LGRAY if ri_p % 2 == 0 else WHITE
        _rut_v   = str(row_p["__rut"]).strip()
        iva_f    = 1.0 if _rut_v in ("CLAY", "-", "", "nan", "None", "NaN") else IVA
        neto_iva = row_p["neto"] * iva_f

        vals = [
            rut,
            nombre if nombre not in ("nan", "", "None") else "(sin nombre)",
            ai.get("descripcion_negocio", "—"),
            ai.get("categoria", "—"),
            ai.get("relevancia_para_cf", "—"),
            row_p["deb"],
            row_p["cre"],
            row_p["neto"],
            neto_iva,
        ]
        for ci, val in enumerate(vals, 1):
            is_iva    = (ci == 9)
            is_text   = ci in (2, 3, 4, 5)   # nombre, descripcion, categoria, relevancia
            is_number = ci >= 6
            c      = ws2.cell(row=row_num, column=ci, value=val)
            c.font = Font(name="Calibri", size=10,
                          color=IVA_TX if is_iva else DARK)
            c.fill = PatternFill("solid",
                                 start_color=IVA_BG if is_iva else bg)
            c.border = _border_thin("DDDDDD")
            if is_number:
                c.number_format = "#,##0;(#,##0);\"-\""
                c.alignment     = Alignment(horizontal="right", vertical="top")
                if not is_iva and ci == 8 and isinstance(val, (int, float)) and val < 0:
                    c.font = Font(name="Calibri", size=10, color=RED)
            elif is_text:
                # Texto largo: justify para que se vea limpio al wrappear
                c.alignment = Alignment(
                    horizontal="justify", vertical="top",
                    wrap_text=True, indent=0
                )
            else:
                # RUT (col 1)
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=False)
        ws2.row_dimensions[row_num].height = 80   # alto suficiente para texto justificado

    r = r + len(grp)

    # Totals row
    for ci in range(1, 6):
        c       = ws2.cell(r, ci, "TOTAL" if ci == 1 else "")
        c.font  = Font(bold=True, name="Calibri", size=10, color=WHITE)
        c.fill  = PatternFill("solid", start_color=NAVY_MID)
        c.border = _border_thin("AAAAAA")
        if ci == 1:
            c.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    for ci in [6, 7, 8]:
        ltr = get_column_letter(ci)
        fc  = ws2.cell(r, ci, f"=SUM({ltr}{tbl_data_start}:{ltr}{r-1})")
        fc.font          = Font(bold=True, name="Calibri", size=10, color=WHITE)
        fc.fill          = PatternFill("solid", start_color=NAVY_MID)
        fc.number_format = "#,##0;(#,##0);\"-\""
        fc.alignment     = Alignment(horizontal="right", vertical="center")
        fc.border        = _border_thin("AAAAAA")
    # IVA total (col 9) — suma por fila con factor correcto según RUT
    iva_total = sum(
        r2["neto"] * (1.0 if str(r2["__rut"]).strip()
                      in ("CLAY", "-", "", "nan", "None", "NaN") else IVA)
        for _, r2 in grp.iterrows()
    )
    c9 = ws2.cell(r, 9, iva_total)
    c9.font          = Font(bold=True, name="Calibri", size=10, color=WHITE)
    c9.fill          = PatternFill("solid", start_color=IVA_TX)
    c9.number_format = "#,##0;(#,##0);\"-\""
    c9.alignment     = Alignment(horizontal="right", vertical="center")
    c9.border        = _border_thin("AAAAAA")
    ws2.row_dimensions[r].height = 22; r += 1

    # IVA note below totals
    ws2.merge_cells(f"A{r}:I{r}")
    c = ws2[f"A{r}"]
    c.value     = "  ★  Proveedores con RUT chileno: valor neto x 1,19 (IVA 19% incluido). Proveedores extranjeros / CLAY (sin RUT): valor neto sin IVA. Solo aplica IVA a gastos nacionales."
    c.font      = Font(name="Calibri", size=9, italic=True, color=IVA_TX)
    c.fill      = PatternFill("solid", start_color=IVA_BG)
    c.alignment = Alignment(indent=1, vertical="center")
    ws2.row_dimensions[r].height = 18; r += 2

    # ── ANÁLISIS COMPLETO ──
    sec_title(r, "📖 Análisis completo"); r += 1
    text_block(r, analysis.get("analisis_general", ""),
               height=160, bg=LGRAY, size=11, color=DARK); r += 2

    # ── POR PERÍODO: tabla + observaciones ──
    multi_period = len(periods) > 1
    if multi_period:
        sec_title(r, "📅 Resumen por período (sin IVA)"); r += 1

        per_hdr = r
        for ci, h in enumerate(["Período", "Mes", "Total cargado", "Total devuelto", "Saldo neto"], 1):
            c           = ws2.cell(per_hdr, ci, h)
            c.font      = Font(bold=True, name="Calibri", size=10, color=WHITE)
            c.fill      = PatternFill("solid", start_color=DGRAY)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = _border_thin("AAAAAA")
        ws2.row_dimensions[per_hdr].height = 24; r += 1

        per_data_start = r
        for pi, period_str in enumerate(periods):
            try:
                p_obj  = pd.Period(period_str)
                mes_es = f"{MONTHS_ES.get(p_obj.month, str(p_obj.month))} {p_obj.year}"
            except Exception:
                mes_es = period_str
            pt     = period_totals.get(period_str, {"deb": 0, "cre": 0})
            bg     = LGRAY if pi % 2 == 0 else WHITE
            neto_p = pt["deb"] - pt["cre"]
            for ci, val in enumerate([period_str, mes_es, pt["deb"], pt["cre"], neto_p], 1):
                c       = ws2.cell(r, ci, val)
                c.font  = Font(name="Calibri", size=10, color=DARK)
                c.fill  = PatternFill("solid", start_color=bg)
                c.border = _border_thin("DDDDDD")
                c.alignment = Alignment(
                    horizontal="right" if ci >= 3 else "center",
                    vertical="center"
                )
                if ci >= 3:
                    c.number_format = "#,##0;(#,##0);\"-\""
                    if ci == 5 and isinstance(val, (int, float)) and val < 0:
                        c.font = Font(name="Calibri", size=10, color=RED)
            ws2.row_dimensions[r].height = 18; r += 1

        for ci in [3, 4, 5]:
            ltr = get_column_letter(ci)
            fc  = ws2.cell(r, ci, f"=SUM({ltr}{per_data_start}:{ltr}{r-1})")
            fc.font          = Font(bold=True, name="Calibri", size=10, color=WHITE)
            fc.fill          = PatternFill("solid", start_color=NAVY_MID)
            fc.number_format = "#,##0;(#,##0);\"-\""
            fc.alignment     = Alignment(horizontal="right", vertical="center")
            fc.border        = _border_thin("AAAAAA")
        for ci in [1, 2]:
            c = ws2.cell(r, ci, "TOTAL" if ci == 1 else "")
            c.font  = Font(bold=True, name="Calibri", size=10, color=WHITE)
            c.fill  = PatternFill("solid", start_color=NAVY_MID)
            c.border = _border_thin("AAAAAA")
        ws2.row_dimensions[r].height = 22; r += 2

        ai_periods = analysis.get("analisis_por_periodo", [])
        if ai_periods:
            sec_title(r, "🔍 Qué pasó cada mes", bg=DGRAY); r += 1
            for ap in ai_periods:
                try:
                    p_obj  = pd.Period(ap.get("periodo", ""))
                    label  = f"{MONTHS_ES.get(p_obj.month, ap.get('periodo',''))} {p_obj.year}"
                except Exception:
                    label  = ap.get("mes_nombre", ap.get("periodo", ""))
                text_block(r, f"  {label}: {ap.get('observacion', '')}",
                           height=28, bg=WHITE, size=10, color=MID); r += 1
        r += 1

    # ── TABLA DISTRIBUCIÓN PROVEEDOR × PERÍODO (CON IVA) ──────────────────────
    # Siempre se muestra (si hay períodos detectados)
    has_periods = len(periods) > 0
    if has_periods and "__period" in df2.columns:
        # Calcular cuántas columnas necesita la tabla dinámica
        n_periods   = len(periods)
        # col 1 = Proveedor, col 2 = RUT, col 3..n+2 = meses, col n+3 = Total c/IVA
        dist_cols   = 2 + n_periods + 1
        dist_last   = get_column_letter(dist_cols)

        # Título con merge dinámico
        ws2.merge_cells(f"A{r}:{dist_last}{r}")
        c = ws2[f"A{r}"]
        c.value     = "DISTRIBUCIÓN DEL GASTO POR PROVEEDOR Y PERÍODO — IVA INCLUIDO SOLO EN PROVEEDORES NACIONALES (CON RUT) ★"
        c.font      = Font(bold=True, name="Calibri", size=10, color=WHITE)
        c.fill      = PatternFill("solid", start_color=NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws2.row_dimensions[r].height = 22; r += 1

        # Subtítulo explicativo
        ws2.merge_cells(f"A{r}:{dist_last}{r}")
        c = ws2[f"A{r}"]
        c.value     = (
            "  ★  Proveedores con RUT chileno: valores incluyen IVA (x1,19). "
            "Proveedores extranjeros / CLAY (sin RUT): valor neto sin IVA."
        )
        c.font      = Font(name="Calibri", size=9, italic=True, color=IVA_TX)
        c.fill      = PatternFill("solid", start_color=IVA_BG)
        c.alignment = Alignment(indent=1, vertical="center")
        ws2.row_dimensions[r].height = 20; r += 1

        # Cabecera de la tabla  — mismo orden que tabla de proveedores: RUT | Nombre
        dist_hdr = r
        # Col 1: RUT
        c = ws2.cell(dist_hdr, 1, "RUT")
        c.font      = Font(bold=True, name="Calibri", size=10, color=WHITE)
        c.fill      = PatternFill("solid", start_color=PERIOD_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = _border_thin("AAAAAA")
        # Col 2: Proveedor
        c = ws2.cell(dist_hdr, 2, "Proveedor")
        c.font      = Font(bold=True, name="Calibri", size=10, color=WHITE)
        c.fill      = PatternFill("solid", start_color=PERIOD_HDR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border_thin("AAAAAA")
        # Cols 3..n+2: meses
        for pi, period_str in enumerate(periods):
            try:
                p_obj  = pd.Period(period_str)
                label  = f"{MONTHS_ES.get(p_obj.month, str(p_obj.month))[:3]}\n{p_obj.year}"
            except Exception:
                label = period_str
            ci = 3 + pi
            c  = ws2.cell(dist_hdr, ci, label)
            c.font      = Font(bold=True, name="Calibri", size=9, color=WHITE)
            c.fill      = PatternFill("solid", start_color=PERIOD_HDR)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border    = _border_thin("AAAAAA")
        # Col total
        ci_total = 3 + n_periods
        c = ws2.cell(dist_hdr, ci_total, "TOTAL\n(IVA si aplica)")
        c.font      = Font(bold=True, name="Calibri", size=10, color=WHITE)
        c.fill      = PatternFill("solid", start_color=TOTAL_COL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = _border_thin("AAAAAA")
        ws2.row_dimensions[dist_hdr].height = 32; r += 1

        # Construir datos de distribución
        # Agrupar por proveedor + período → neto (deb - cre) para que cuadre con tabla proveedores
        has_period_col = "__period" in df2.columns
        if has_period_col:
            grp_dist = df2[mask].copy()
            grp_dist["__period_str"] = grp_dist["__period"].astype(str)
            pivot_data = grp_dist.groupby(
                ["__rut", "__nombre", "__period_str"], as_index=False
            ).agg(deb=("__deb", "sum"), cre=("__cre", "sum"))
            pivot_data["neto"] = pivot_data["deb"] - pivot_data["cre"]

            # Sin proveedor también
            if len(df_no_rut) > 0 and df_no_rut["__deb"].sum() > 0:
                norut_dist = df2[~mask].copy()
                norut_dist["__period_str"] = norut_dist["__period"].astype(str)
                norut_grp  = norut_dist.groupby("__period_str", as_index=False).agg(
                    deb=("__deb", "sum"), cre=("__cre", "sum"))
                norut_grp["neto"]     = norut_grp["deb"] - norut_grp["cre"]
                norut_grp["__rut"]    = "-"
                norut_grp["__nombre"] = "Sin proveedor identificado"
                pivot_data = pd.concat(
                    [pivot_data, norut_grp[["__rut", "__nombre", "__period_str", "neto"]]],
                    ignore_index=True
                )

            # Pivot: filas = proveedor, columnas = periodo → usando neto para cuadrar con tabla superior
            pivot = pivot_data.pivot_table(
                index=["__rut", "__nombre"],
                columns="__period_str",
                values="neto",
                aggfunc="sum",
                fill_value=0
            )
            # Aplicar IVA por fila: solo proveedores con RUT chileno
            for _idx_row in pivot.index:
                _rut_v2 = str(_idx_row[0]).strip()
                _iva_f2 = 1.0 if _rut_v2 in ("CLAY", "-", "", "nan", "None", "NaN") else IVA
                pivot.loc[_idx_row, :] = pivot.loc[_idx_row, :] * _iva_f2
            # Ordenar columnas por período
            period_cols_avail = [p for p in periods if p in pivot.columns]
            pivot = pivot.reindex(columns=period_cols_avail, fill_value=0)
            # Ordenar filas por total descendente
            pivot["__total"] = pivot.sum(axis=1)
            pivot = pivot.sort_values("__total", ascending=False)

            # Rellenar filas de la hoja
            dist_data_start = r
            for pi_row, (idx, prow) in enumerate(pivot.iterrows()):
                rut_v    = idx[0]
                nombre_v = idx[1]
                bg       = LGRAY if pi_row % 2 == 0 else WHITE

                # Col 1: RUT  — mismo orden que tabla de proveedores arriba
                c = ws2.cell(r, 1, rut_v)
                c.font      = Font(name="Calibri", size=9, color=MID)
                c.fill      = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = _border_thin("DDDDDD")
                # Col 2: Nombre
                c = ws2.cell(r, 2, nombre_v if nombre_v not in ("nan","","None") else "(sin nombre)")
                c.font      = Font(name="Calibri", size=10, color=DARK)
                c.fill      = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(horizontal="justify", wrap_text=True, vertical="center")
                c.border    = _border_thin("DDDDDD")
                # Cols de períodos
                for pi2, period_str in enumerate(periods):
                    ci2   = 3 + pi2
                    val_p = float(prow.get(period_str, 0))
                    c     = ws2.cell(r, ci2, val_p if val_p != 0 else "")
                    c.font      = Font(name="Calibri", size=10, color=DARK)
                    c.fill      = PatternFill("solid", start_color=bg)
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    c.border    = _border_thin("DDDDDD")
                    if val_p != 0:
                        c.number_format = "#,##0;(#,##0);\"-\""
                # Total c/IVA
                total_v = float(prow["__total"])
                ct = ws2.cell(r, ci_total, total_v)
                ct.font          = Font(bold=True, name="Calibri", size=10, color=IVA_TX)
                ct.fill          = PatternFill("solid", start_color=IVA_BG)
                ct.number_format = "#,##0;(#,##0);\"-\""
                ct.alignment     = Alignment(horizontal="right", vertical="center")
                ct.border        = _border_thin("AAAAAA")
                ws2.row_dimensions[r].height = 20; r += 1

            # Fila de totales por período
            for ci2 in range(1, dist_cols + 1):
                ltr2 = get_column_letter(ci2)
                if ci2 <= 2:
                    c = ws2.cell(r, ci2, "TOTAL" if ci2 == 1 else "")
                    c.font  = Font(bold=True, name="Calibri", size=10, color=WHITE)
                    c.fill  = PatternFill("solid", start_color=TOTAL_COL)
                    c.border = _border_thin("AAAAAA")
                else:
                    fc = ws2.cell(r, ci2,
                        f"=SUM({ltr2}{dist_data_start}:{ltr2}{r-1})")
                    fc.font          = Font(bold=True, name="Calibri", size=10, color=WHITE)
                    fc.fill          = PatternFill("solid",
                                          start_color=TOTAL_COL if ci2 == ci_total else PERIOD_HDR)
                    fc.number_format = "#,##0;(#,##0);\"-\""
                    fc.alignment     = Alignment(horizontal="right", vertical="center")
                    fc.border        = _border_thin("AAAAAA")
            ws2.row_dimensions[r].height = 22; r += 2

    # ── COLUMN WIDTHS ──
    # A-I: usados por tabla de proveedores
    # C en adelante: también usados por tabla de distribución (períodos)
    # Se usa un ancho que funcione para AMBOS usos (texto wrapeado + números monetarios)
    ws2.column_dimensions["A"].width = 14   # RUT
    ws2.column_dimensions["B"].width = 32   # Nombre proveedor
    ws2.column_dimensions["C"].width = 36   # Descripción / Período 1
    ws2.column_dimensions["D"].width = 20   # Categoría  / Período 2
    ws2.column_dimensions["E"].width = 36   # Relevancia / Período 3
    ws2.column_dimensions["F"].width = 16   # Total cargado / Período 4
    ws2.column_dimensions["G"].width = 16   # Total abonado / Período 5
    ws2.column_dimensions["H"].width = 16   # Saldo neto    / Período 6
    ws2.column_dimensions["I"].width = 18   # Neto + IVA    / Período 7

    # Anchos uniformes para columnas de períodos en tabla de distribución.
    # Se aplican AL FINAL para garantizar que todos los meses tengan exactamente el mismo ancho.
    # Usar 22 porque es lo suficientemente ancho para montos grandes Y para texto corto.
    if has_periods and "__period" in df2.columns:
        COL_PERIOD_W = 22   # ancho uniforme para cada columna de mes
        COL_TOTAL_W  = 22   # columna TOTAL c/IVA (mismo ancho, se distingue por color)
        _n_periods   = len(periods)
        _ci_total    = 2 + _n_periods + 1
        for _pi in range(_n_periods):
            ws2.column_dimensions[get_column_letter(3 + _pi)].width = COL_PERIOD_W
        ws2.column_dimensions[get_column_letter(_ci_total)].width = COL_TOTAL_W

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# ── Normalización de nombres de proveedores con IA ───────────────────────────
def normalizar_proveedores_con_ia(nombres: list, api_key: str) -> dict:
    """
    Usa la IA para unificar nombres que corresponden a la misma empresa o persona,
    sin importar capitalización, idioma, abreviación o variante del nombre.
    Retorna dict {nombre_original: nombre_canonico}
    """
    if not nombres or not api_key:
        return {}

    client = Groq(api_key=api_key)

    prompt = f"""Eres un experto en identificar empresas y personas en registros contables chilenos.
Tengo una lista de nombres de proveedores. Tu tarea es identificar cuáles se refieren
a la MISMA empresa o persona, aunque estén escritos diferente.

CRITERIO: Usa tu conocimiento de negocios para identificar si son la misma entidad.
No te limites a comparar texto — razona si comercialmente son lo mismo.

EJEMPLOS DE LO QUE DEBES UNIFICAR:
  - "facebook" / "Facebook" / "FACEBOOK" / "FACEBK"   → "Facebook"
  - "Google Ads" / "GOOGLE ADS" / "google ads"
    / "Google Anuncios"                                → "Google Ads"
  - "LinkedIn" / "LINKEDIN" / "Linkedin"              → "LinkedIn"
  - "Juan Perez" / "JUAN PEREZ" / "juan perez"        → "Juan Perez"
  - "X Corp Paid Features" / "Corp Paid Features"
    / "Twitter Ads"                                    → "X / Twitter"
  - "Spotify" / "SPOTIFY" / "Spotify AB"              → "Spotify"
  - Mismo nombre con/sin tildes o con sigla obvia     → unificar

SOLO NO UNIFICAR cuando sean claramente empresas distintas
(diferente razón social, diferente rubro, diferente persona).

LISTA DE NOMBRES A PROCESAR ({len(nombres)} en total):
{json.dumps(nombres, ensure_ascii=False, indent=2)}

TAREA:
1. Identifica grupos de nombres que son la misma entidad
2. Elige el nombre canónico más reconocible y en Title Case para cada grupo
3. Mapea TODOS los nombres originales a su canónico (incluso los sin duplicados)

Responde ÚNICAMENTE con JSON válido, sin explicaciones, sin comentarios:
{{"nombre_original_exacto_de_la_lista": "Nombre Canonico"}}"""

    try:
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[
                {"role": "system",
                 "content": (
                     "Eres un experto en resolución de entidades comerciales. "
                     "Identificas cuándo dos nombres distintos se refieren a la misma "
                     "empresa o persona usando conocimiento de negocios, no solo texto."
                 )},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            max_tokens=2048,
        )
        text = response.choices[0].message.content.strip()
        if "```json" in text:
            text = text.split("```json")[1].split("```")[0].strip()
        elif "```" in text:
            text = text.split("```")[1].split("```")[0].strip()
        return _parse_json_safe(text)
    except Exception:
        return {}   # si falla, continuar sin normalizar


# ── Clay enrichment (Playwright headless) ───────────────────────────────────
def _extraer_proveedor_de_detalle(texto: str) -> str:
    """
    Extrae el nombre del proveedor/concepto del texto descriptivo del asiento en Clay.
    Ejemplos:
      "Gastos Varios# 202602 del mes de Febrero de 2026 Facebook"  → "Facebook"
      "Pago Honorarios# 202603 del mes de Marzo de 2026 Juan Perez" → "Juan Perez"
    """
    if not texto or texto.strip() in ("", "-"):
        return "(sin descripcion)"
    texto = texto.strip()
    # Patrón más común: "... de YYYY <Vendor>"
    match = re.search(r"de \d{4}\s+(.+)$", texto, re.IGNORECASE | re.DOTALL)
    if match:
        vendor = match.group(1).strip()
        if vendor:
            return vendor[:80]
    # Fallback: último segmento después de guion/barra
    parts = re.split(r"\s[-–/]\s", texto)
    if len(parts) > 1:
        candidate = parts[-1].strip()
        if len(candidate) > 2:
            return candidate[:80]
    return texto[:80]


def buscar_detalles_clay(asientos: list, clay_email: str, clay_password: str,
                         status_fn=None) -> dict:
    """
    Para una lista de números de asiento contable (sin RUT), entra a Clay en modo
    headless, busca cada asiento y extrae la descripción para identificar el proveedor.

    Retorna: dict {str(asiento_num): vendor_name}
    """
    try:
        from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
    except ImportError:
        return {}   # playwright no instalado — omitir enriquecimiento

    resultados = {}
    if not asientos or not clay_email or not clay_password:
        return resultados

    # Selectores Clay
    SEL_SEARCH   = "#search-input"
    SEL_RESULT_A = (
        "#content > div > div > div.col.s12.l12 > ul > li "
        "> div.col.s12.l12.title_asiento > div.accounting_title "
        "> div.info_acconting > div > span:nth-child(2) > a"
    )
    SEL_DETAIL = (
        "#transaction-details > div.col.s12.l12.view_details "
        "> div.box-scroll-mobile.col.l12.s12 > p"
    )

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=True)
        context = browser.new_context()
        page    = context.new_page()

        try:
            # ── LOGIN ──────────────────────────────────────────────────────────
            if status_fn:
                status_fn("🔑 Iniciando sesión en Clay (headless)...")
            page.goto("https://app.clay.cl/login", wait_until="domcontentloaded")
            page.wait_for_selector("#email", timeout=30000)
            page.click("#email")
            page.type("#email", clay_email, delay=50)
            page.click("#password")
            page.type("#password", clay_password, delay=50)
            page.click("#submit-login")
            page.wait_for_timeout(9000)

            # ── NAVEGAR A ASIENTOS (URL directa) ───────────────────────────────
            if status_fn:
                status_fn("📂 Abriendo Gestión Contable → Asientos...")
            page.goto(
                "https://app.clay.cl/accountingentries/timeline",
                wait_until="domcontentloaded",
            )
            page.wait_for_timeout(3000)

            # ── ITERAR ASIENTOS ────────────────────────────────────────────────
            for i, asiento_num in enumerate(asientos):
                asiento_str = str(asiento_num).strip()
                if asiento_str in ("", "nan", "None", "NaN"):
                    continue
                if status_fn:
                    status_fn(
                        f"🔍 Asiento {asiento_str}  ({i+1}/{len(asientos)}) — "
                        "buscando en Clay..."
                    )
                try:
                    # Limpiar campo y escribir "numero:XXXXX"
                    search = page.locator(SEL_SEARCH)
                    search.wait_for(timeout=10000)
                    search.click()
                    page.keyboard.press("Control+A")
                    page.keyboard.press("Backspace")
                    page.type(SEL_SEARCH, f"numero:{asiento_str}", delay=30)
                    page.keyboard.press("Enter")
                    page.wait_for_timeout(3500)

                    # Esperar resultado y hacer click (abre nueva pestaña)
                    link = page.locator(SEL_RESULT_A).first
                    link.wait_for(state="visible", timeout=8000)

                    with context.expect_page() as popup_info:
                        link.click()
                    popup = popup_info.value
                    popup.wait_for_load_state("domcontentloaded", timeout=15000)
                    popup.wait_for_timeout(2000)

                    # Leer texto descriptivo del asiento
                    try:
                        detail_text = (
                            popup.locator(SEL_DETAIL)
                                 .inner_text(timeout=8000)
                                 .strip()
                        )
                        vendor = _extraer_proveedor_de_detalle(detail_text)
                        resultados[asiento_str] = vendor
                    except Exception:
                        resultados[asiento_str] = "(detalle no disponible)"
                    finally:
                        try:
                            popup.close()
                        except Exception:
                            pass

                except Exception:
                    resultados[asiento_str] = "(no encontrado)"
                    # Cerrar posibles pestanas huerfanas
                    for extra in context.pages[1:]:
                        try:
                            extra.close()
                        except Exception:
                            pass
                    # Volver a asientos si se perdio la navegacion
                    try:
                        if "accountingentries" not in page.url:
                            page.goto(
                                "https://app.clay.cl/accountingentries/timeline",
                                wait_until="domcontentloaded",
                            )
                            page.wait_for_timeout(2000)
                    except Exception:
                        pass

        except Exception as e:
            if status_fn:
                status_fn(f"⚠️ Error Clay: {e}")
        finally:
            try:
                browser.close()
            except Exception:
                pass

    return resultados


# ── UI ────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### ⚙️ Configuración")
    api_key = st.text_input(
        "API Key de Groq (gratis)",
        type="password",
        value=os.getenv("GROQ_API_KEY", ""),
        help="100% gratis en console.groq.com → API Keys → Create API Key",
    )
    st.divider()
    st.markdown("""
    ### ℹ️ ¿Cómo usar?
    1. Sube tu Libro Mayor (.xlsx)
    2. Haz clic en **Analizar con IA**
    3. Revisa el análisis en pantalla
    4. Descarga tu informe Excel
    """)
    st.divider()
    st.caption("ComunidadFeliz SPA · Agente Contable IA · v2.1")

st.title("📊 Agente Analizador de Libro Mayor")
st.markdown(
    "Sube tu reporte contable y la IA **investiga cada proveedor**, "
    "analiza los gastos y genera un **informe Excel completo** explicado de forma simple."
)
st.divider()

uploaded = st.file_uploader(
    "📁 Sube tu archivo Libro Mayor (.xlsx)",
    type=["xlsx"],
    help="Formato estándar del sistema contable de Comunidad Feliz",
)

if uploaded:
    st.success(f"✅ Archivo cargado: **{uploaded.name}**")

    col_l, col_c, col_r = st.columns([1, 2, 1])
    with col_c:
        btn = st.button(
            "🚀 Analizar con IA",
            type="primary",
            use_container_width=True,
            disabled=not api_key,
        )

    if not api_key:
        st.warning("⚠️ Ingresa tu API Key de Groq en el panel lateral. Consíguela gratis en console.groq.com")

    if btn and api_key:
        prog = st.progress(0)

        try:
            prog.progress(10, "📂 Leyendo el Libro Mayor...")
            df, title, company, saldo_ant, account_name, periods, period_totals = \
                parse_libro_mayor(uploaded)

            n_tx         = len(df)
            multi_period = len(periods) > 1
            period_info  = f" · {len(periods)} meses detectados" if multi_period else ""
            prog.progress(25, f"✅ {n_tx} transacciones — cuenta: {account_name}{period_info}")

            if multi_period:
                st.info(
                    f"📅 **Reporte multi-período:** {len(periods)} meses — "
                    + ", ".join(periods)
                    + ". El informe incluirá desglose por mes **con IVA**."
                )

            with st.expander(f"📋 Vista previa — {n_tx} transacciones", expanded=False):
                display_df = df[[c for c in df.columns if not c.startswith("__")]]
                st.dataframe(display_df, use_container_width=True, height=280)

            deb_col      = df.columns[3] if len(df.columns) > 3 else None
            cre_col      = df.columns[4] if len(df.columns) > 4 else None
            asiento_col  = df.columns[2] if len(df.columns) > 2 else None
            rut_col      = df.columns[7] if len(df.columns) > 7 else None
            nombre_col   = df.columns[8] if len(df.columns) > 8 else None

            # ── ENRIQUECIMIENTO CLAY ──────────────────────────────────────────
            # Si hay credenciales de Clay y existen filas sin RUT, consultamos
            # cada asiento en Clay para identificar el proveedor o concepto real.
            clay_email    = os.getenv("CLAY_EMAIL", "")
            clay_password = os.getenv("CLAY_PASSWORD", "")
            clay_enriquecidos = 0

            if clay_email and clay_password and rut_col and asiento_col and nombre_col:
                _no_rut_mask = df[rut_col].apply(
                    lambda x: str(x).strip() in ("", "nan", "NaN", "-", "None")
                )
                if _no_rut_mask.sum() > 0:
                    _asientos_sin_rut = (
                        df.loc[_no_rut_mask, asiento_col]
                        .dropna()
                        .astype(str)
                        .str.strip()
                        .unique()
                        .tolist()
                    )
                    _asientos_sin_rut = [
                        a for a in _asientos_sin_rut
                        if a not in ("", "nan", "None", "NaN")
                    ]

                    if _asientos_sin_rut:
                        _clay_status = st.empty()
                        prog.progress(30, f"🔍 Consultando Clay para {len(_asientos_sin_rut)} asientos sin proveedor...")

                        def _update_clay(msg):
                            _clay_status.info(f"Clay · {msg}")

                        clay_map = buscar_detalles_clay(
                            _asientos_sin_rut,
                            clay_email,
                            clay_password,
                            status_fn=_update_clay,
                        )
                        _clay_status.empty()

                        # Actualizar df: asignar vendor encontrado como "contraparte"
                        # y usar RUT sintético "CLAY" para que aparezca como proveedor propio
                        for idx, row in df[_no_rut_mask].iterrows():
                            asiento_v = str(row[asiento_col]).strip()
                            vendor    = clay_map.get(asiento_v, "")
                            if vendor and not vendor.startswith("("):
                                df.at[idx, rut_col]    = "CLAY"
                                df.at[idx, nombre_col] = vendor
                                clay_enriquecidos += 1

                        if clay_enriquecidos:
                            st.success(
                                f"✅ Clay identificó **{clay_enriquecidos} transacciones** "
                                f"sin proveedor en {len(_asientos_sin_rut)} asientos consultados."
                            )
                        else:
                            st.warning(
                                "🔍 Clay fue consultado pero no pudo identificar proveedores "
                                "en los asientos sin RUT. Quedan como 'Sin proveedor identificado'."
                            )
            # ─────────────────────────────────────────────────────────────────

            # ── NORMALIZACIÓN DE NOMBRES ──────────────────────────────────────
            # Unifica variantes del mismo proveedor antes de agrupar y analizar.
            # Trabaja sobre TODOS los nombres (con RUT oficial + CLAY + sin RUT).
            if nombre_col:
                _todos_nombres = (
                    df[nombre_col]
                    .dropna()
                    .astype(str)
                    .str.strip()
                    .unique()
                    .tolist()
                )
                _todos_nombres = [
                    n for n in _todos_nombres
                    if n not in ("", "nan", "None", "NaN", "-",
                                 "Transacciones sin proveedor identificado",
                                 "Gastos Varios")
                ]
                if len(_todos_nombres) > 1:
                    prog.progress(37, "🔤 Normalizando nombres de proveedores...")
                    _nombre_map = normalizar_proveedores_con_ia(_todos_nombres, api_key)
                    if _nombre_map:
                        df[nombre_col] = df[nombre_col].apply(
                            lambda x: _nombre_map.get(
                                str(x).strip(),
                                str(x).strip()
                            )
                        )
            # ─────────────────────────────────────────────────────────────────

            providers = []
            if rut_col and nombre_col:
                for _, row in df[[rut_col, nombre_col]].drop_duplicates().iterrows():
                    rut    = str(row[rut_col])
                    nombre = str(row[nombre_col])
                    if rut not in ("nan", "-", "", "None") and nombre not in ("nan", "", "None"):
                        providers.append({"rut": rut, "nombre": nombre})

            total_deb = to_num(df[deb_col]).sum() if deb_col else 0
            total_cre = to_num(df[cre_col]).sum() if cre_col else 0
            neto      = total_deb - total_cre

            period_summary = ""
            if multi_period:
                lines = []
                for p in periods:
                    pt = period_totals.get(p, {"deb": 0, "cre": 0})
                    lines.append(f"  · {p}: cargado ${pt['deb']:,.0f} / devuelto ${pt['cre']:,.0f}")
                period_summary = "\nDETALLE POR PERÍODO:\n" + "\n".join(lines)

            tx_summary = (
                f"- Cuenta analizada: {account_name}\n"
                f"- Total transacciones: {n_tx}\n"
                f"- Total cargado (débitos): ${total_deb:,.0f}\n"
                f"- Total devuelto (créditos): ${total_cre:,.0f}\n"
                f"- Saldo neto: ${neto:,.0f}\n"
                f"- Período(s): {title}\n"
                f"- Meses con movimientos: {', '.join(periods) if periods else 'N/A'}\n"
                f"- Proveedores con RUT identificado: {len(providers)}"
                + period_summary
            )

            prog.progress(40, f"🤖 Analizando {len(providers)} proveedores con IA...")

            analysis = get_ai_analysis(providers, tx_summary, account_name, api_key, periods)

            prog.progress(75, "📊 Generando informe Excel...")

            excel_bytes = generate_excel(
                df, analysis, title, company, saldo_ant,
                account_name, periods, period_totals
            )

            prog.progress(100, "✅ ¡Listo!")

            st.divider()
            st.markdown(f"### 📌 Cuenta: `{account_name}`")

            m1, m2, m3, m4 = st.columns(4)
            m1.metric("Transacciones",  n_tx)
            m2.metric("Total cargado",  f"${total_deb:,.0f}")
            m3.metric("Total devuelto", f"${total_cre:,.0f}")
            m4.metric("Proveedores",    len(providers))

            st.info(
                f"💡 **IVA:** El neto contable es **${neto:,.0f}**. "
                "El Excel aplica IVA 19% **solo a proveedores con RUT chileno** (operan en Chile). "
                "Proveedores extranjeros / CLAY (sin RUT) quedan con su valor neto sin IVA. "
                "Ver columna 'Neto + IVA' en el informe."
            )

            with st.expander("📌 Resumen ejecutivo", expanded=True):
                st.info(analysis.get("resumen_ejecutivo", ""))

            if analysis.get("alertas"):
                with st.expander("⚡ Observaciones", expanded=True):
                    for a in analysis["alertas"]:
                        st.warning(f"· {a}")

            if multi_period and analysis.get("analisis_por_periodo"):
                with st.expander("📅 Análisis por mes", expanded=True):
                    for ap in analysis["analisis_por_periodo"]:
                        try:
                            p_obj  = pd.Period(ap.get("periodo", ""))
                            label  = f"**{MONTHS_ES.get(p_obj.month, ap.get('periodo',''))} {p_obj.year}**"
                        except Exception:
                            label = f"**{ap.get('mes_nombre', ap.get('periodo', ''))}**"
                        st.markdown(f"{label}: {ap.get('observacion', '')}")

            if multi_period and period_totals:
                with st.expander("📅 Resumen por período (con IVA)", expanded=False):
                    per_rows = []
                    for p in periods:
                        try:
                            p_obj  = pd.Period(p)
                            mes    = f"{MONTHS_ES.get(p_obj.month, p)} {p_obj.year}"
                        except Exception:
                            mes = p
                        pt   = period_totals[p]
                        neto_p = pt["deb"] - pt["cre"]
                        per_rows.append({
                            "Período":      p,
                            "Mes":          mes,
                            "Cargado ($)":  f"${pt['deb']:,.0f}",
                            "Devuelto ($)": f"${pt['cre']:,.0f}",
                            "Neto ($)":     f"${neto_p:,.0f}",
                        })
                    st.dataframe(pd.DataFrame(per_rows), use_container_width=True, hide_index=True)

            with st.expander("📖 Análisis completo", expanded=False):
                st.write(analysis.get("analisis_general", ""))

            st.divider()
            fname = f"Analisis_{account_name[:20].replace(' ','_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
            st.download_button(
                label="📥 Descargar informe Excel",
                data=excel_bytes,
                file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                type="primary",
                use_container_width=True,
            )
            st.success(f"✅ Informe listo: **{fname}**")

        except json.JSONDecodeError as e:
            prog.empty()
            st.error(f"Error procesando respuesta de IA: {e}. Intenta de nuevo.")
        except Exception as e:
            prog.empty()
            st.error(f"Error: {e}")

else:
    st.markdown(
        """
        <div style='text-align:center;padding:50px 20px;background:white;
             border-radius:8px;border:1px solid #ddd;margin-top:20px;'>
            <h3 style='color:#333;margin:0 0 10px'>Sube tu Libro Mayor</h3>
            <p style='color:#777;font-size:15px;margin:0'>
                Arrastra tu archivo .xlsx aquí o usa el botón de arriba
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )
